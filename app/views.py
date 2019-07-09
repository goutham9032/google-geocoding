# Python imports
import time
import json
import urllib
import re

from itertools import zip_longest

# Django imports
from django.shortcuts import render
from django.conf import settings
from django.http import HttpResponse, JsonResponse

# 3rd party imports
import requests
import openpyxl
import xlwt


def get_rows_from_exel(file_location):
    wb_obj = openpyxl.load_workbook(file_location)
    sheet = wb_obj.active
    max_row = sheet.max_row
    max_column = sheet.max_column
    rows = [
        sheet.cell(row=i, column=j).value
        for i in range(1, max_row + 1)
        for j in range(1, max_column + 1)
    ]
    return rows


def upload_exel(request):
    if request.method == "GET":
        return render(request, "home.html", {"a": 1})
    exel_obj = request.FILES["exel_file"]
    file_name = re.sub("[^a-zA-Z0-9\n\.]", "-", exel_obj.name)
    file_id = "%s_%s" % (int(time.time()), file_name)
    file_location = "/tmp/%s" % (file_id)
    with open(file_location, "wb+") as f:
        for chunk in exel_obj.chunks():
            f.write(chunk)
    rows = get_rows_from_exel(file_location)
    api_url = "https://maps.googleapis.com/maps/api/geocode/json?"
    get_params = lambda x: urllib.parse.urlencode(
        {"address": x, "key": settings.GEOCODING_API_KEY}
    )
    req = requests.Session()
    lat_long_lists = []
    for address in rows:
        map_url = api_url + get_params(address)
        resp = req.get(url=map_url).json()
        results = resp["results"][0].get("geometry", "").get("location", "")
        lat, lng = results["lat"], results["lng"]
        lat_long_lists.append([address, lat, lng])
    rows = tuple(map(tuple, lat_long_lists))
    book = openpyxl.Workbook()
    sheet = book.active
    for row in rows:
        sheet.append(row)

    book.save(file_location.replace("/tmp/", "/tmp/results_"))
    download_url = "/download_results/%s" % (file_id.split(".")[0])
    return JsonResponse({"success": True, "download_url": download_url})


def download_results(request, file_id):
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="results.xls"'
    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Lat_Long_results")

    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ["Address", "latitude", "Longitude"]

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()

    file_path = "/tmp/results_%s.xlsx" % (file_id)
    results = get_rows_from_exel(file_path)
    results = list(zip_longest(*[iter(results)] * 3))
    for row in results:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
    wb.save(response)
    return response
